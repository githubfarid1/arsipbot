from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from openpyxl import Workbook, load_workbook
import setting as s
import os
import time
import sys
import json
import argparse

def browser_init():
    # WARNING: AGAR BISA SET LOKASI DOWNLOAD PDF, MAKA NAMA PROFILE HARUS "DEFAULT" 
    cud = s.CHROME_USER_DATA
    cp = s.CHROME_PROFILE
    options = webdriver.ChromeOptions()
    options.add_argument("user-data-dir={}".format(cud))
    options.add_argument("profile-directory={}".format(cp))
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    options.add_argument("--window-size=800,600")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    profile = {"download.default_directory": s.CHROME_DOWNLOAD_PATH + os.sep, 
                "download.extensions_to_open": "applications/pdf",
                "download.prompt_for_download": False,
                'profile.default_content_setting_values.automatic_downloads': 1,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True                   
                }
    options.add_experimental_option("prefs", profile)

    return webdriver.Chrome(service=Service(executable_path=os.path.join(os.getcwd(), "chromedriver", "chromedriver.exe")), options=options)

def parse(xlsfile):
    # breakpoint()
    # user = 'bwsmalukuutara'
    # passwd = 'P@sswd2022!'
    url = 'https://arsip-sda.pusair-pu.go.id/admin/dashboard/'
    # xlsfile = 'data_arsip_tata_tahun_penataan_2024.xlsx'
    
    workbook = load_workbook(filename=xlsfile, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    # breakpoint()
    firstrow = 10
    datalist= []
    actnoberkas = worksheet['A{}'.format(firstrow)].value
    actnobox = worksheet['K{}'.format(firstrow)].value
    actthcipta = worksheet['F{}'.format(firstrow)].value
    acttitle = worksheet['D{}'.format(firstrow)].value
    actklas = worksheet['C{}'.format(firstrow)].value

    for i in range(firstrow, worksheet.max_row):
        # breakpoint()
        noberkas = worksheet['A{}'.format(i)].value
        thcipta = worksheet['F{}'.format(i)].value
        title = worksheet['D{}'.format(i)].value
        nobox = worksheet['K{}'.format(i)].value
        klas = worksheet['C{}'.format(i)].value

        if nobox == None:
            nobox = actnobox
        else:
            if actnobox != nobox:
                actnobox = nobox

        if noberkas == None:
            noberkas = actnoberkas
            thcipta = actthcipta
            title = acttitle
            klas = actklas
        else:
            if actnoberkas != noberkas:
                actnoberkas = noberkas
                actthcipta = thcipta 
                acttitle = title
                actklas = klas
                # breakpoint()

        mdict = {
        "noberkas": noberkas,
        "noitem": worksheet['B{}'.format(i)].value,
        "thcipta": thcipta,
        "thtata": worksheet.title,
        "klasifikasi": klas,
        "nobox": nobox,
        "title": title,
        "uraian": worksheet['E{}'.format(i)].value,
        "keamanan": worksheet['L{}'.format(i)].value,
        "jenisarsip": 'Dinamis',
        "ket": 'COPY',
        "jumlah":worksheet['H{}'.format(i)].value,
        "bentukarsip": 'Buku',
        "rak": '1'}
        datalist.append(mdict)
    # breakpoint()
    # sys.exit()

    driver = browser_init()
    driver.get(url)
    driver.maximize_window()
    try:
        driver.find_element(By.CSS_SELECTOR, "input[name='login']").clear()
        driver.find_element(By.CSS_SELECTOR, "input[name='password']").clear()    
        driver.find_element(By.CSS_SELECTOR, "input[name='login']").send_keys(s.USER)
        driver.find_element(By.CSS_SELECTOR, "input[name='password']").send_keys(s.PASSWORD)    
        driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    except:
        pass

    # breakpoint()
    for data in datalist:
        # breakpoint()
        driver.get("https://arsip-sda.pusair-pu.go.id/admin/archive/add")
        # time.sleep(2)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "input[name='file_num']")))
        driver.find_element(By.CSS_SELECTOR, "input[name='file_num']").send_keys(data['noberkas'])
        driver.find_element(By.CSS_SELECTOR, "input[name='item_num']").send_keys(data['noitem'])
        driver.find_element(By.CSS_SELECTOR, "input[name='year_file']").send_keys(data['thcipta'])        
        driver.find_element(By.CSS_SELECTOR, "input[name='year_archive']").clear()
        driver.find_element(By.CSS_SELECTOR, "input[name='year_archive']").send_keys(data['thtata'])
        # breakpoint()
        driver.find_elements(By.CSS_SELECTOR, "span[class='select2-selection__rendered']")[1].click()
        driver.find_element(By.CSS_SELECTOR, "input[class='select2-search__field']").send_keys(data['klasifikasi'])
        time.sleep(1)
        driver.find_element(By.CSS_SELECTOR, "li[class='select2-results__option select2-results__option--highlighted']").click()
        driver.find_elements(By.CSS_SELECTOR, "span[class='select2-selection__rendered']")[2].click()
        # breakpoint()
        driver.find_element(By.CSS_SELECTOR, "input[class='select2-search__field']").send_keys(f"{data['nobox']} - Rak/Lemari {data['rak']}({data['thtata']})")
        time.sleep(1)
        driver.find_element(By.CSS_SELECTOR, "li[class='select2-results__option select2-results__option--highlighted']").click()
        driver.find_element(By.CSS_SELECTOR, "input[name='document_name']").send_keys(data['title'])
        driver.find_element(By.CSS_SELECTOR, "textarea[name='document_note']").send_keys(data['uraian'])
        
        Select(driver.find_element(By.CSS_SELECTOR, "select[name='daftar_archive']")).select_by_visible_text(data['keamanan'])
        
        Select(driver.find_element(By.CSS_SELECTOR, "select[name='archive_type']")).select_by_visible_text(data['jenisarsip'])
        Select(driver.find_element(By.CSS_SELECTOR, "select[name='satuan']")).select_by_visible_text(data['bentukarsip'])
        driver.find_element(By.CSS_SELECTOR, "input[name='total']").send_keys(data['jumlah'])
        driver.find_element(By.CSS_SELECTOR, "input[id='inline-{}']".format(data['ket'])).click()
        # breakpoint()
        submit = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']")))
        # breakpoint()
        try:
            submit.click()
        except:
            submit.click()
        time.sleep(1)

def main():
    parser = argparse.ArgumentParser(description="ARSIP BOT")
    parser.add_argument('-i', '--input', type=str,help="File Input")
    args = parser.parse_args()
    if not args.input:
        print('use: python arsipbot.py -i <filename>')
        exit()

    if args.input[-5:] != '.xlsx':
        print('File input have to XLSX file')
        exit()
    
    parse(xlsfile=args.input)
    print("End Process...")
if __name__ == '__main__':
    main()
